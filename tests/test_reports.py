import json
from dataclasses import dataclass

from openpyxl import load_workbook
from pdfminer.high_level import extract_text

from app.repositories.reports.db_report import DataBaseReport
from app.data.models.groups import Group
from app.data.models.students import Student
from app.data.models.subjects import Subject
from app.data.models.timetables import TimeTable
from app.repositories.reports.json_report import JsonReport
from app.repositories.reports.pdf_report import PDFReport
from app.repositories.reports.xlsx_report import XLSXReport
from app.data.models.base_class import Base


@dataclass
class MockStudent:
    full_name: str = 'John'
    age: int = 20
    gender: str = 'M'
    weight: int = 90
    height: int = 180
    average_score: float = 3.5


@dataclass
class MockSubject:
    name: str = 'Math'
    day_of_week: str = 'Mon'
    time: str = '07:00'


@dataclass
class MockTimeTable:
    subjects: list[MockSubject]
    week: tuple[str, ...]
    times: tuple[str, ...]


@dataclass
class MockGroup:
    students: list[MockStudent]
    timetable: MockTimeTable
    name: str = 'MyGroup'


mock_timetable = MockTimeTable([MockSubject(), MockSubject(), MockSubject()],
                               ('Mon',), ('07:00',))

mock_timetable2 = MockTimeTable([MockSubject('Info'), MockSubject(name='Info'), MockSubject('Info')],
                                ('Mon',), ('07:00',))

mock_groups = [
    MockGroup([MockStudent(), MockStudent(), MockStudent()], mock_timetable),
    MockGroup([MockStudent(), MockStudent(), MockStudent()], mock_timetable2)
]


def test_xlsx_report(tmpdir):
    file_path = tmpdir.join('report')

    getter = XLSXReport(mock_groups, file_path)
    getter.get_report()

    wb = load_workbook(f'{file_path}.xlsx')
    assert wb.sheetnames == ['Группа MyGroup', 'Группа MyGroup1']
    assert len(list(wb.active)) == 4

    for sheet in wb.worksheets:
        for row in list(sheet)[1:]:
            assert [cell.value for cell in row] == ['John', 20, 'M', 90, 180, 3.5]


test_json_data = {
    'Группа MyGroup': {
        'Студент John': {
            "Возраст": 20,
            "Пол": 'M',
            "Вес": 90,
            "Рост": 180,
            "Средний балл": 3.5,
        },
        'Студент John1': {
            "Возраст": 20,
            "Пол": 'M',
            "Вес": 90,
            "Рост": 180,
            "Средний балл": 3.5,
        },
        'Студент John2': {
            "Возраст": 20,
            "Пол": 'M',
            "Вес": 90,
            "Рост": 180,
            "Средний балл": 3.5,
        },
    },
    'Группа MyGroup1': {
        'Студент John': {
            "Возраст": 20,
            "Пол": 'M',
            "Вес": 90,
            "Рост": 180,
            "Средний балл": 3.5,
        },
        'Студент John1': {
            "Возраст": 20,
            "Пол": 'M',
            "Вес": 90,
            "Рост": 180,
            "Средний балл": 3.5,
        },
        'Студент John2': {
            "Возраст": 20,
            "Пол": 'M',
            "Вес": 90,
            "Рост": 180,
            "Средний балл": 3.5,
        },
    },
}


def test_json_report(tmpdir):
    file_path = tmpdir.join('report')

    getter = JsonReport(mock_groups, file_path)
    getter.get_report()

    with open(f'{file_path}.json', 'r') as file:
        data = json.load(file)

    assert data == test_json_data


test_pdf_data = ['Группа MyGroup',
                 '- John - Возраст: 20, Пол: M, Рост: 180, Вес: 90, Средний балл: 3.5',
                 '- John - Возраст: 20, Пол: M, Рост: 180, Вес: 90, Средний балл: 3.5',
                 '- John - Возраст: 20, Пол: M, Рост: 180, Вес: 90, Средний балл: 3.5',
                 'Группа MyGroup',
                 '- John - Возраст: 20, Пол: M, Рост: 180, Вес: 90, Средний балл: 3.5',
                 '- John - Возраст: 20, Пол: M, Рост: 180, Вес: 90, Средний балл: 3.5',
                 '- John - Возраст: 20, Пол: M, Рост: 180, Вес: 90, Средний балл: 3.5']


def test_pdf_report(tmpdir):
    file_path = tmpdir.join('report')

    getter = PDFReport(mock_groups, file_path)
    getter.get_report()

    report_data = extract_text(f'{file_path}.pdf').strip().split('\n')
    report_data = [i.strip() for i in report_data if i]

    assert report_data == test_pdf_data


def test_db_report():
    engine = "postgresql://postgres:postgres@localhost/test"

    getter = DataBaseReport(mock_groups, engine=engine)
    Base.metadata.create_all(bind=getter.engine)

    session = getter.session

    session.query(Student).delete()
    session.query(Subject).delete()
    session.query(TimeTable).delete()
    session.query(Group).delete()
    session.commit()

    getter.get_report()

    groups_name = [i[0] for i in session.query(Group.name)]
    students_name = [i[0] for i in session.query(Student.full_name).join(Group)]
    groups_subj = [i[0] for i in session.query(Subject.name).join(TimeTable).join(Group)]

    assert groups_name == ['MyGroup', 'MyGroup']
    assert students_name == ['John', 'John', 'John', 'John', 'John', 'John']
    assert groups_subj == ['Math', 'Math', 'Math', 'Info', 'Info', 'Info']
