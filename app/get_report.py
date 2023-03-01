from abc import ABC, abstractmethod

from entities import Group


class IReportGetter(ABC):
    def __init__(self, groups: list[Group], report_path: str = "report"):
        self.groups = groups
        self.report_path = report_path

    @abstractmethod
    def get_report(self):
        pass


class XLSXReport(IReportGetter):
    def get_report(self) -> None:
        from openpyxl.workbook import Workbook

        wb = Workbook()
        ws = wb.active
        wb.remove(ws)
        for group in self.groups:
            ws = wb.create_sheet(f"Группа {group.name}")
            ws.append(
                [
                    "ФИО",
                    "Возраст",
                    "Пол",
                    "Вес",
                    "Рост",
                    "Средний балл",
                ]
            )
            for student in group.students:
                ws.append(list(vars(student).values()))

        wb.save(f"{self.report_path}.xlsx")


class JsonReport(IReportGetter):
    def get_report(self) -> None:
        import json

        report = {}
        for group in self.groups:
            report[f"Group {group.name}"] = {}
            for student in group.students:
                student_data = list(vars(student).values())[1:]
                report[f"Group {group.name}"][f"Student {student.full_name}"] = dict(
                    zip(
                        [
                            "Возраст",
                            "Пол",
                            "Вес",
                            "Рост",
                            "Средний балл",
                        ],
                        student_data,
                    )
                )
        with open(f"{self.report_path}.json", "w", encoding="utf8") as file:
            json.dump(report, file, indent=4, ensure_ascii=False)


class PDFReport(IReportGetter):
    def get_report(self):
        from fpdf import FPDF

        pdf = FPDF()
        pdf.add_font("DejaVu", "", "app/fonts_for_pdf/DejaVuSansCondensed.ttf", uni=True)
        pdf.add_font("DejaVu", "B", "app/fonts_for_pdf/DejaVuSansCondensed-Bold.ttf", uni=True)

        for group in self.groups:
            pdf.add_page()
            pdf.set_font("DejaVu", size=25, style="B")
            pdf.cell(400, 15, txt=f"Группа {group.name}", ln=1)

            for student in group.students:
                pdf.set_font("DejaVu", size=10, style="")
                pdf.cell(
                    400,
                    8,
                    txt=f"- {student.full_name} - Возраст: {student.age}, Пол: {student.gender},"
                    f" Рост: {student.height}, Вес: {student.weight}, Средний балл: {student.average_score}",
                    ln=1,
                )

        pdf.output(f"{self.report_path}.pdf")
